"""File parsing utilities: PDF, DOCX, XLSX → text chunks + image extraction for vision.

Uses PyMuPDF (fitz) as the primary PDF engine for reliable text extraction
and image rendering. Every PDF page always produces a chunk (even image-only
pages) so page counts are always correct.
"""

from __future__ import annotations

import base64
import io
import logging
from pathlib import Path

logger = logging.getLogger(__name__)


# ── PDF Parsing ──────────────────────────────────────────────────────────────

def parse_pdf(filepath: str) -> list[dict[str, str]]:
    """Extract text from each page of a PDF using PyMuPDF.

    CRITICAL: Every page produces a chunk, even if no text is found.
    This ensures page counts are always accurate and the vision agent
    can process image-only engineering drawings.
    """
    chunks: list[dict[str, str]] = []
    source_name = Path(filepath).name

    # Primary: PyMuPDF (fitz) — handles most PDF types reliably
    try:
        import fitz  # PyMuPDF

        doc = fitz.open(filepath)
        total_pages = len(doc)

        for i, page in enumerate(doc):
            text = page.get_text("text") or ""
            text = text.strip()

            if text and len(text) > 20:
                # Page has meaningful extractable text
                chunks.append({
                    "source": source_name,
                    "page": str(i + 1),
                    "content": text,
                    "has_text": "true",
                })
            else:
                # Image-only page (typical for engineering drawings)
                # Still create a chunk so page count is correct and
                # the vision agent knows to analyze this page
                chunks.append({
                    "source": source_name,
                    "page": str(i + 1),
                    "content": (
                        f"[Engineering drawing — Page {i + 1} of {total_pages}. "
                        f"This page contains graphical/visual content with minimal or no "
                        f"extractable text. Visual inspection by the Vision Agent is required "
                        f"for QA/QC review of this page.]"
                    ),
                    "has_text": "false",
                })

        doc.close()
        logger.info(
            "Parsed %d pages from %s via PyMuPDF (%d with text, %d image-only)",
            total_pages, source_name,
            sum(1 for c in chunks if c.get("has_text") == "true"),
            sum(1 for c in chunks if c.get("has_text") == "false"),
        )
        return chunks

    except ImportError:
        logger.warning("PyMuPDF not installed, falling back to pdfplumber")
    except Exception as exc:
        logger.error("PyMuPDF parse error for %s: %s, falling back to pdfplumber", filepath, exc)

    # Fallback: pdfplumber
    try:
        import pdfplumber

        with pdfplumber.open(filepath) as pdf:
            total_pages = len(pdf.pages)
            for i, page in enumerate(pdf.pages):
                text = page.extract_text() or ""
                text = text.strip()

                if text and len(text) > 20:
                    chunks.append({
                        "source": source_name,
                        "page": str(i + 1),
                        "content": text,
                        "has_text": "true",
                    })
                else:
                    chunks.append({
                        "source": source_name,
                        "page": str(i + 1),
                        "content": (
                            f"[Engineering drawing — Page {i + 1} of {total_pages}. "
                            f"Image-only page, visual inspection required.]"
                        ),
                        "has_text": "false",
                    })
    except ImportError:
        logger.warning("Neither PyMuPDF nor pdfplumber installed; returning empty chunks")
    except Exception as exc:
        logger.error("PDF parse error for %s: %s", filepath, exc)

    return chunks


def get_pdf_page_count(filepath: str) -> int:
    """Get the total number of pages in a PDF without full parsing."""
    try:
        import fitz
        doc = fitz.open(filepath)
        count = len(doc)
        doc.close()
        return count
    except Exception:
        pass

    try:
        import pdfplumber
        with pdfplumber.open(filepath) as pdf:
            return len(pdf.pages)
    except Exception:
        pass

    return 0


def extract_pdf_page_images(filepath: str, dpi: int = 200) -> list[dict[str, str]]:
    """Convert each PDF page to a base64-encoded PNG image for vision analysis.

    Uses PyMuPDF (fitz) as the primary renderer. Falls back to pdf2image.
    Returns list of dicts with keys: source, page, image_b64.
    """
    images: list[dict[str, str]] = []
    source_name = Path(filepath).name

    # Primary: PyMuPDF (no external binary dependency)
    try:
        import fitz

        doc = fitz.open(filepath)
        for i, page in enumerate(doc):
            mat = fitz.Matrix(dpi / 72, dpi / 72)
            pix = page.get_pixmap(matrix=mat, alpha=False)
            img_bytes = pix.tobytes("png")
            b64 = base64.b64encode(img_bytes).decode("utf-8")
            images.append({
                "source": source_name,
                "page": str(i + 1),
                "image_b64": b64,
            })
        doc.close()
        logger.info("Extracted %d page images from %s via PyMuPDF", len(images), source_name)
        return images
    except ImportError:
        pass
    except Exception as exc:
        logger.warning("PyMuPDF image extraction failed for %s: %s, trying pdf2image", filepath, exc)

    # Fallback: pdf2image (requires poppler binaries)
    try:
        from pdf2image import convert_from_path

        pil_images = convert_from_path(filepath, dpi=dpi)
        for i, img in enumerate(pil_images):
            buffer = io.BytesIO()
            img.save(buffer, format="PNG")
            b64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
            images.append({
                "source": source_name,
                "page": str(i + 1),
                "image_b64": b64,
            })
        logger.info("Extracted %d page images from %s via pdf2image", len(images), source_name)
        return images
    except ImportError:
        logger.warning("Neither PyMuPDF nor pdf2image available for image extraction")
    except Exception as exc:
        logger.error("pdf2image extraction failed for %s: %s", filepath, exc)

    return images


# ── DOCX Parsing ─────────────────────────────────────────────────────────────

def parse_docx(filepath: str) -> list[dict[str, str]]:
    """Extract text from a DOCX file, one chunk per heading-section."""
    try:
        from docx import Document
    except ImportError:
        logger.warning("python-docx not installed; returning empty chunks")
        return []

    chunks: list[dict[str, str]] = []
    try:
        doc = Document(filepath)
        current_heading = "General"
        current_text: list[str] = []

        for para in doc.paragraphs:
            if para.style.name.startswith("Heading"):
                if current_text:
                    chunks.append({
                        "source": Path(filepath).name,
                        "section": current_heading,
                        "content": "\n".join(current_text),
                    })
                current_heading = para.text or "Untitled Section"
                current_text = []
            else:
                if para.text.strip():
                    current_text.append(para.text)

        if current_text:
            chunks.append({
                "source": Path(filepath).name,
                "section": current_heading,
                "content": "\n".join(current_text),
            })
    except Exception as exc:
        logger.error("DOCX parse error for %s: %s", filepath, exc)
    return chunks


# ── XLSX Parsing ─────────────────────────────────────────────────────────────

def parse_xlsx(filepath: str) -> list[dict[str, str]]:
    """Extract data from each sheet of an XLSX file as tabular text."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl not installed; returning empty chunks")
        return []

    chunks: list[dict[str, str]] = []
    try:
        wb = load_workbook(filepath, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_text: list[str] = []
            for row in ws.iter_rows(values_only=True):
                row_str = " | ".join(str(c) if c is not None else "" for c in row)
                if row_str.strip(" |"):
                    rows_text.append(row_str)
            if rows_text:
                chunks.append({
                    "source": Path(filepath).name,
                    "sheet": sheet_name,
                    "content": "\n".join(rows_text),
                })
    except Exception as exc:
        logger.error("XLSX parse error for %s: %s", filepath, exc)
    return chunks


# ── Router ───────────────────────────────────────────────────────────────────

def parse_file(filepath: str) -> list[dict[str, str]]:
    """Route to the correct parser based on file extension."""
    ext = Path(filepath).suffix.lower()
    if ext == ".pdf":
        return parse_pdf(filepath)
    elif ext == ".docx":
        return parse_docx(filepath)
    elif ext == ".xlsx":
        return parse_xlsx(filepath)
    elif ext in (".png", ".jpg", ".jpeg", ".tiff", ".tif", ".bmp"):
        # Standalone image files (floor plans, site sketches, scanned drawings)
        return _parse_image(filepath)
    elif ext in (".dwg", ".dxf"):
        logger.warning("CAD file detected (%s): %s — native parsing not supported, "
                       "convert to PDF for analysis", ext, filepath)
        return [{
            "source": Path(filepath).name,
            "section": "CAD Notice",
            "content": (
                f"[CAD file ({ext}) detected. Native DWG/DXF parsing is not supported. "
                f"Please export/convert this file to PDF format for AI-powered review. "
                f"Most CAD software (AutoCAD, Revit, BricsCAD) can export to PDF.]"
            ),
            "has_text": "false",
        }]
    else:
        logger.warning("Unsupported file type: %s", ext)
        return []


def _parse_image(filepath: str) -> list[dict[str, str]]:
    """Parse a standalone image file (PNG, JPG, TIFF) as a single-page drawing."""
    source_name = Path(filepath).name
    return [{
        "source": source_name,
        "page": "1",
        "content": (
            f"[Standalone image file: {source_name}. "
            f"This is a scanned drawing/floor plan/site sketch. "
            f"Visual inspection by the OCR and Vision Agents is required.]"
        ),
        "has_text": "false",
    }]


def extract_image_as_b64(filepath: str) -> list[dict[str, str]]:
    """Convert a standalone image file to base64 for vision analysis."""
    images: list[dict[str, str]] = []
    source_name = Path(filepath).name
    try:
        with open(filepath, "rb") as f:
            raw = f.read()
        b64 = base64.b64encode(raw).decode("utf-8")
        images.append({
            "source": source_name,
            "page": "1",
            "image_b64": b64,
        })
        logger.info("Loaded image %s as base64 (%d KB)", source_name, len(raw) // 1024)
    except Exception as exc:
        logger.error("Failed to read image %s: %s", filepath, exc)
    return images

